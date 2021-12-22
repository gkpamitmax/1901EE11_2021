from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from fpdf import FPDF 
import datetime
import shutil
file1 = None
file2 = None
file3 = None
stamp = ""
signature = ""
generate_all_transcript=False


def marksheet_generator():

   data={}
   subject_data={}
   Roll_data={}

   folder="media"
   if not os.path.exists(folder):
        return
   grades=os.path.join(folder,file1.name)
   names_Roll=os.path.join(folder,file2.name)
   subjects_master=os.path.join(folder,file3.name)

   DIR = "marksheet"
   if os.path.exists(DIR):
      shutil.rmtree(DIR)
   os.makedirs(DIR)
   
 # Opening the subject_master and store in a subject_data
   with open(subjects_master,'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
          subject_data[row['subno']]=[row['subname'],row['ltp'],row['crd']]

 # Opening the grades and store in a data     
   with open(grades, 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
         x=(row['Roll'],row['Sem'])
         if x in data:
               data[x] +=[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]
         else:
             data[x] =[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]
 #Opening the names_Roll and store in Roll_data
   with open(names_Roll, 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
         Roll_data[row['Roll']]=row['Name']

 # Grades Map
   grades = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7,
          'CC': 6, 'CD': 5, 'DD': 4,'DD*':4,'F*':0, 'F': 0, 'I': 0}     
              
 # now making the overall sheet for all Roll no
   with open(names_Roll, 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
          Roll_path = os.path.join(DIR, row['Roll']+".xlsx")
          wb = Workbook()
          wb.create_sheet(index=0,title="Overall")
          sheet=wb["Overall"]
          sheet.append(["Roll No.",row['Roll']])
          sheet.append(["Name of Student ",row['Name']])
          a=row['Roll']
          sheet.append(["Discipline",a[4:6]])
          
          n=[]
          spi=[]
          credit_sem=[]
          total_credit=[]
          sem_credit_cleared=[]
          cpi=[]
          x=0
          w=0
          for j in range(1,11):
               v=(a,str(j))
               if v in data:
                     count=0
                     credit_obtain=0
                     credit_cleared=0
                     for i in data[v]:
                       count+=int(i[1])
                       credit_obtain+=grades[i[3]]*(int(i[1]))
                       if grades[i[3]]>0:
                           credit_cleared+=int(i[1])
               if v in data:
                  n.append(j)   
                  credit_sem.append(count)     
                  spi.append(credit_obtain/count)
                  sem_credit_cleared.append(credit_cleared)
                  x+=count
                  total_credit.append(x)
                  w+=credit_obtain
                  cpi.append(w/x)
          sheet.append(["Semester No."]+[  x for x in n])
          sheet.append(["Semester wise Credit Taken"]+[x for x in credit_sem])
          sheet.append(["SPI"]+[round(x,2) for x in spi])
          sheet.append(["Total Credits Taken"]+[x for x in total_credit])
          sheet.append(["CPI"]+[round(x,2) for x in cpi])
          sheet.append(["credit_cleared"]+[x for x in sem_credit_cleared])
          wb.remove(wb["Sheet"])
          wb.save(Roll_path)  
  # load the workbook create extra sheet and add the semester detail         
          wb=load_workbook(Roll_path)
          p = 0
          for j in range(1,11):
               v=(a,str(j))
               if v in data:
                    p=p+1
                    wb.create_sheet(index=p,title="Sem{}".format(str(j)))
                    wb["Sem{}".format(str(j))].append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                    sheet=wb["Sem{}".format(str(j))]
                    for i in data[v]:
                         x=subject_data[i[0]]
                         sheet.append([ sheet.max_row,i[0],x[0],x[1],(x[2]),i[2],i[3] ])
          wb.save(Roll_path)

def transcript_create_range(range_start,range_end):
  output_DIR1=r"./marksheet"
  
  absent=[]
  class PDF(FPDF):
    pass # nothing happens when it is executed. 
  for x in range(int(range_start[6:8]),int(range_end[6:8])+1):
    y=''
    z=''
    if x>=1 and x<10:
       z=range_start[0:6]+'0'+str(x)
       y=range_start[0:6]+'0'+str(x)+".xlsx"
    else:
        z=range_start[0:6]+str(x)
        y=range_start[0:6]+str(x)+".xlsx"
    y=output_DIR1+'/'+y

    if os.path.exists(y):
    
     
        xlsx = openpyxl.load_workbook(y)
       
        sheet = xlsx.active

    
        total_semester=len(xlsx.sheetnames)-1
        
        #student_information
        stud_name=sheet.cell(row=2, column=2).value
        year_of_joining="20"+range_start[0:2]
        course=sheet.cell(row=3, column=2).value
        programe_taken=""
        if(range_start[2:4]=="01"):
            programe_taken= "Bachelor of Technology"
        elif(range_start[2:4]=="21"):
            programe_taken= "Doctor of Philosphy"
        else:
            programe_taken= "Master of Technology"
        
      
         
        pdf=PDF("L","mm","A3") if programe_taken == "Bachelor of Technology" else PDF("L","mm","A4") #page format. A4 is the default value of the format, you don't have to specify it.
        pdf.add_page()

        # Set font: Arial size font
        font=8 if programe_taken == "Bachelor of Technology" else 5.5
        pdf.set_font('Arial', size=font)
        xstart=4
        ystart=5
        xend=430  if programe_taken == "Bachelor of Technology" else 305
        yend=287  if programe_taken == "Bachelor of Technology" else 210

        pdf.image("./image.jpg",xstart,ystart,xend,yend)

        
        sem_info=[]            
        sem_courses_info=[]   
        for sem in range(1,total_semester+1):
            sem_info.append([sheet.cell(row=5, column=1+sem).value,sheet.cell(row=6, column=1+sem).value,sheet.cell(row=8, column=1+sem).value,sheet.cell(row=9, column=1+sem).value])
        for sem in range(1,total_semester+1):
            sheet = xlsx.worksheets[sem]
            sem_course=[]
            sem_course.append(["Sub Code","Subject Name","L-T-P","CRD","GRD"])
            for course_count in range(1,sheet.max_row):
                sem_course.append([sheet.cell(row=1+course_count, column=2).value,sheet.cell(row=1+course_count, column=3).value,sheet.cell(row=1+course_count, column=4).value,sheet.cell(row=1+course_count, column=5).value,sheet.cell(row=1+course_count, column=7).value])
            sem_courses_info.append(sem_course)
        
        
        #setting font for student info
        pdf.set_font_size(font)
        if programe_taken != "Bachelor of Technology":
            #setting Roll no
            pdf.set_xy(87.0,33.5)
            pdf.cell(w=15,h=5,txt=x,border=0)
            
            #setting name of student
            pdf.set_xy(130.0,33.5)
            pdf.cell(w=5,h=5,txt=stud_name,border=0)
            
            #setting year of admission
            pdf.set_xy(225.0,33.5)
            pdf.cell(w=10,h=5,txt=year_of_joining,border=0)

            #setting programe_taken
            pdf.set_xy(87.5,38)
            pdf.cell(w=20,h=5,txt=programe_taken,border=0)

            #setting programe_taken
            pdf.set_xy(130,38)
            pdf.cell(w=25,h=5,txt=course,border=0)
        else:
            pdf.set_font_size(font+2)
            #setting Roll_no
            pdf.set_xy(120,45.6)
            pdf.cell(w=15,h=5,txt=z,border=0)
            
            #setting name of student
            pdf.set_xy(180.0,45.6)
            pdf.cell(w=5,h=5,txt=stud_name,border=0)
            
            #setting year of admission
            pdf.set_xy(315.0,45.6)
            pdf.cell(w=10,h=5,txt=year_of_joining,border=0)

            #setting programe_taken
            pdf.set_xy(120,51.4)
            pdf.cell(w=20,h=5,txt=programe_taken,border=0)

            #setting course
            pdf.set_xy(180,51.4)
            pdf.cell(w=25,h=5,txt=course,border=0)

            pdf.set_font_size(font)


      
        sem_count=0
   
        rowh=4 if programe_taken == "Bachelor of Technology" else 3
        space_x=8 if programe_taken == "Bachelor of Technology" else 6
        x=20  if programe_taken == "Bachelor of Technology" else 17
        y=45
        
        
        if programe_taken == "Bachelor of Technology":
            y=60
        ymax=0
       
        epw = pdf.w - 2*pdf.l_margin
        sem_width=epw/3.3                   
        for sem in sem_courses_info:
            sem_count+=1
            if(sem_count==4 or sem_count==7):
                y=ymax+10

            if sem_count==1 or sem_count==4 or sem_count==7:
                pdf.set_xy(x,y)
            else:
                pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),y)
            
            pdf.set_font('Arial','U')
            pdf.cell(x,rowh,"Semester "+str(sem_count))
            pdf.set_font('Arial')
            if sem_count>1:
                if sem_count==4 or sem_count==7:
                    pdf.set_xy(x+sem_width*((sem_count-1)%3),y+rowh+2)
                else:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),y+rowh+2)
            else:
                pdf.set_xy(x,y+rowh+2)
    
            col_width = [0.18*sem_width,0.55*sem_width,0.12*sem_width, 0.075*sem_width, 0.075*sem_width]
            
            #start of sem
            for row in sem:
                col=0
                for datum in row:
                    pdf.cell(col_width[col], rowh, str(datum), border=1,align='C')
                    col+=1
                if sem_count!=1 and sem_count!=4 and sem_count!=7:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),pdf.get_y()+rowh)
                else:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),pdf.get_y()+rowh)
                ymax=max(pdf.get_y(),ymax)

            pdf.set_xy(pdf.get_x(),pdf.get_y()+2)
            credits_taken=sem_info[sem_count-1][0]
            credits_cleared=sem_info[sem_count-1][3]
            spi=sem_info[sem_count-1][1]
            cpi=sem_info[sem_count-1][2]
            sem_text="Credits Taken: "+ str(credits_taken)+"     Credits Cleared:  "+ str(credits_cleared)+"      SPI: "+str(spi)+"     CPI: "+str(cpi)
            pdf.cell(sem_width*0.85, rowh,sem_text , border=1,align='C')
            if(sem_count==total_semester or sem_count==3 or sem_count==6 or sem_count==9):
                if sem_count>=1 and sem_count<=3:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                elif sem_count>=4 and sem_count<=7:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                else:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)



        
        if programe_taken == "Bachelor of Technology" and stamp != "":
            pdf.image(stamp,145,225,32,32)
        elif (stamp != ""):
            pdf.image(stamp,100,178,20,20)
        if programe_taken == "Bachelor of Technology" and signature != "":
            pdf.image(signature,325,235,20,20)
        elif (signature != ""):
            pdf.image(signature,230,172,12,12)
        pdf.set_xy(15,187)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(20,255)
        pdf.set_font("Arial","",10)
        now = datetime.datetime.now()
        
        pdf.cell(w=0,h=2,txt="Date Generated: "+(now.strftime("%d %B %Y, %H:%M")) ,border=0)
        pdf.set_xy(220,183)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(320,255)
        pdf.cell(w=0,h=2,txt=" ______________________",border=0)
        pdf.set_xy(220,187)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(320,260)
        pdf.cell(w=0,h=1,txt="Asistant Registrar(Academic)",border=0)
        
        #saving the genearated pdf
        pdf.output(f'transcriptsIITP\\{z}.pdf','F')
    else:
        absent.append(z)
  return absent

def transcript_create():
    output_DIR1=r"./marksheet"
    
    class PDF(FPDF):
        pass # nothing happens when it is executed. 
    
    for file_Roll in os.listdir(output_DIR1):
        Roll=file_Roll[0:file_Roll.index('.')]
        file_Roll=output_DIR1+"/"+file_Roll
        
        xlsx = openpyxl.load_workbook(file_Roll)

        sheet = xlsx.active

        #finding total semester by no of sheets in a xlxs
        total_semester=len(xlsx.sheetnames)-1
        
        #student information
        stud_name=sheet.cell(row=2, column=2).value
        year_of_joining="20"+Roll[0:2]
        programe_taken=""
        if(Roll[2:4]=="01"):programe_taken= "Bachelor of Technology"
        elif(Roll[2:4]=="21"):programe_taken= "Doctor of Philosphy"
        else:programe_taken= "Master of Technology"
        course=sheet.cell(row=3, column=2).value

         
        pdf=PDF("L","mm","A3") if programe_taken == "Bachelor of Technology" else PDF("L","mm","A4") #page format. A4 is the default value of the format, you don't have to specify it.
        pdf.add_page()

        # Set font: Arial size font
        font=8 if programe_taken == "Bachelor of Technology" else 5.5
        pdf.set_font('Arial', size=font)
        xstart=4
        ystart=5
        xend=430  if programe_taken == "Bachelor of Technology" else 305
        yend=287  if programe_taken == "Bachelor of Technology" else 210

        pdf.image("./image.jpg",xstart,ystart,xend,yend)

        sem_info=[]            
        sem_courses_info=[]     
        for sem in range(total_semester+1):
            if sem==0:continue
            sem_info.append([sheet.cell(row=5, column=1+sem).value,sheet.cell(row=6, column=1+sem).value,sheet.cell(row=8, column=1+sem).value,sheet.cell(row=9, column=1+sem).value])
        for sem in range(total_semester+1):
            if(sem==0):continue
            sheet = xlsx.worksheets[sem]
            sem_course=[]
            sem_course.append(["Sub Code","Subject Name","L-T-P","CRD","GRD"])
            for course_count in range(sheet.max_row):
                if(course_count==0):continue
                sem_course.append([sheet.cell(row=1+course_count, column=2).value,sheet.cell(row=1+course_count, column=3).value,sheet.cell(row=1+course_count, column=4).value,sheet.cell(row=1+course_count, column=5).value,sheet.cell(row=1+course_count, column=7).value])
            sem_courses_info.append(sem_course)
        
        
        #setting font for student info
        pdf.set_font_size(font)
        if programe_taken != "Bachelor of Technology":
            #setting Roll no
            pdf.set_xy(87.0,33.5)
            pdf.cell(w=15,h=5,txt=Roll,border=0)
            
            #setting name of student
            pdf.set_xy(130.0,33.5)
            pdf.cell(w=5,h=5,txt=stud_name,border=0)
            
            #setting year of admission
            pdf.set_xy(225.0,33.5)
            pdf.cell(w=10,h=5,txt=year_of_joining,border=0)

            #setting programe_taken
            pdf.set_xy(87.5,38)
            pdf.cell(w=20,h=5,txt=programe_taken,border=0)

            #setting programe_taken
            pdf.set_xy(130,38)
            pdf.cell(w=25,h=5,txt=course,border=0)
        else:
            pdf.set_font_size(font+2)
            #setting Roll no
            pdf.set_xy(120,45.6)
            pdf.cell(w=15,h=5,txt=Roll,border=0)
            
            #setting name of student
            pdf.set_xy(180.0,45.6)
            pdf.cell(w=5,h=5,txt=stud_name,border=0)
            
            #setting year of admission
            pdf.set_xy(315.0,45.6)
            pdf.cell(w=10,h=5,txt=year_of_joining,border=0)

            #setting programe_taken
            pdf.set_xy(120,51.4)
            pdf.cell(w=20,h=5,txt=programe_taken,border=0)

            #setting course
            pdf.set_xy(180,51.4)
            pdf.cell(w=25,h=5,txt=course,border=0)

            pdf.set_font_size(font)


        #logic for building semester tables
        sem_count=0
        #row_height
        rowh=4 if programe_taken == "Bachelor of Technology" else 3
        space_x=8 if programe_taken == "Bachelor of Technology" else 6
        x=20  if programe_taken == "Bachelor of Technology" else 17
        y=45
        
        
        if programe_taken == "Bachelor of Technology":
            y=60
        ymax=0
        # Effective page width, or just epw
        epw = pdf.w - 2*pdf.l_margin
        sem_width=epw/3.3                    # total 3 semester in a row : 0.3 added to consider space for gaps between 2 consecutive sem in a row
        for sem in sem_courses_info:
            sem_count+=1
            if(sem_count==4 or sem_count==7):
                y=ymax+10

            if sem_count==1 or sem_count==4 or sem_count==7:
                pdf.set_xy(x,y)
            else:
                pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),y)
            
            pdf.set_font('Arial','U')
            pdf.cell(x,rowh,"Semester "+str(sem_count))
            pdf.set_font('Arial')
            if sem_count>1:
                if sem_count==4 or sem_count==7:
                    pdf.set_xy(x+sem_width*((sem_count-1)%3),y+rowh+2)
                else:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),y+rowh+2)
            else:
                pdf.set_xy(x,y+rowh+2)
    
            col_width = [0.18*sem_width,0.55*sem_width,0.12*sem_width, 0.075*sem_width, 0.075*sem_width]
            
            #start of sem
            for row in sem:
                col=0
                for datum in row:
                    pdf.cell(col_width[col], rowh, str(datum), border=1,align='C')
                    col+=1
                if sem_count!=1 and sem_count!=4 and sem_count!=7:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),pdf.get_y()+rowh)
                else:
                    pdf.set_xy(x+(sem_width+space_x)*((sem_count-1)%3),pdf.get_y()+rowh)
                ymax=max(pdf.get_y(),ymax)

            pdf.set_xy(pdf.get_x(),pdf.get_y()+2)
            credits_taken=sem_info[sem_count-1][0]
            credits_cleared=sem_info[sem_count-1][3]
            spi=sem_info[sem_count-1][1]
            cpi=sem_info[sem_count-1][2]
            sem_text="Credits Taken: "+ str(credits_taken)+"     Credits Cleared:  "+ str(credits_cleared)+"      SPI: "+str(spi)+"     CPI: "+str(cpi)
            pdf.cell(sem_width*0.85, rowh,sem_text , border=1,align='C')
            if(sem_count==total_semester or sem_count==3 or sem_count==6 or sem_count==9):
                if sem_count>=1 and sem_count<=3:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                elif sem_count>=4 and sem_count<=7:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                else:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)



        
        if programe_taken == "Bachelor of Technology" and stamp != "":
            pdf.image(stamp,145,225,32,32)
        elif (stamp != ""):
            pdf.image(stamp,100,178,20,20)
        if programe_taken == "Bachelor of Technology" and signature != "":
            pdf.image(signature,325,235,20,20)
        elif (signature != ""):
            pdf.image(signature,230,172,12,12)
        pdf.set_xy(15,187)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(20,255)
        pdf.set_font("Arial","",10)
        now = datetime.datetime.now()
        
        pdf.cell(w=0,h=2,txt="Date Generated: "+(now.strftime("%d %B %Y, %H:%M")) ,border=0)
        pdf.set_xy(220,183)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(320,255)
        pdf.cell(w=0,h=2,txt=" ______________________",border=0)
        pdf.set_xy(220,187)
        if programe_taken == "Bachelor of Technology":
            pdf.set_xy(320,260)
        pdf.cell(w=0,h=1,txt="Asistant Registrar(Academic)",border=0)

        #saving the genearated pdf
        pdf.output(f'transcriptsIITP\\{Roll}.pdf','F')
    return
# Create your views here.
def home(request):
    if request.method == "POST":

        output_DIR10=r"./media"
        if os.path.exists(output_DIR10):
            shutil.rmtree(output_DIR10)

        output_DIR=r"./transcriptsIITP"
        if os.path.exists(output_DIR):
            shutil.rmtree(output_DIR)
        
        global file1
        file1 = request.FILES['file1']
        fs1 = FileSystemStorage()
        fs1.save(file1.name,file1)
        global file2
        file2 = request.FILES['file2']
        fs2 = FileSystemStorage()
        fs2.save(file2.name,file2)
        global file3
        file3 = request.FILES['file3']
        fs3 = FileSystemStorage()
        fs3.save(file3.name,file3)        
        if 'seal' in request.FILES.keys():
            uploaded_file4 = request.FILES['seal']
            fs4 = FileSystemStorage()
            fs4.save(uploaded_file4.name,uploaded_file4)
            global stamp
            stamp = os.path.join('media',uploaded_file4.name)
        if 'sign' in request.FILES.keys():
            uploaded_file5 = request.FILES['sign']
            fs5 = FileSystemStorage()
            fs5.save(uploaded_file5.name,uploaded_file5)
            global signature
            signature = os.path.join('media',uploaded_file5.name)


        return render(request,'myapp/index.html')
    return render(request,'myapp/index.html')

def range_generator(request):
    if request.method == "POST":
        output_DIR=r"./transcriptsIITP"
        if os.path.exists(output_DIR):
            shutil.rmtree(output_DIR)
        os.makedirs(output_DIR)
        range_start = request.POST.get('start')
        range_end = request.POST.get('end')
        range_start = range_start.upper()
        range_end = range_end.upper()
        marksheet_generator()
        absent=[]
        absent=transcript_create_range(range_start,range_end)

        output_DIR2=r"./marksheet"
        if os.path.exists(output_DIR2):
           shutil.rmtree(output_DIR2)

        absent_list = {'absent':absent}

        return render(request,'myapp/index.html',absent_list)

    return render(request,'myapp/index.html')

def generate_all_roll_transcript(request):

    output_DIR=r"./transcriptsIITP"
    if os.path.exists(output_DIR):
        shutil.rmtree(output_DIR)
    os.makedirs(output_DIR)
    marksheet_generator()
    transcript_create()
    output_DIR2=r"./marksheet"
    if os.path.exists(output_DIR2):
        shutil.rmtree(output_DIR2)
    

    return render(request,'myapp/index.html')