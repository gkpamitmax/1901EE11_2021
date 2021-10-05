def generate_marksheet():
    

# Importing Required Libraries
   import csv
   import os
   import shutil
   from openpyxl import Workbook
   from openpyxl import load_workbook
   from openpyxl.reader.excel import load_workbook
   os.system('cls')

    # Check if Grades Folder Already Exists, If yes Delete and finally Remake
   if(os.path.isdir(r'.\output')):
       shutil.rmtree('.\output')
   os.makedirs('.\output')

   #using dictionary we map all 3 input file in dictionary data,subject_data,roll_data
   data={}
   subject_data={}
   roll_data={}
 # Opening the subject_master and store in a subject_data
   with open('subjects_master.csv','r') as file:
      reader = csv.DictReader(file)
      for row in reader:
          subject_data[row['subno']]=[row['subname'],row['ltp'],row['crd']]
 # Opening the grades and store in a data     
   with open('grades.csv', 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
         x=(row['Roll'],row['Sem'])
         if x in data:
               data[x] +=[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]
         else:
             data[x] =[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]
 #Opening the subject_master and store in a subject_data type container
   with open('names-roll.csv', 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
         roll_data[row['Roll']]=row['Name']

 # Grades Map
   grades = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7,
          'CC': 6, 'CD': 5, 'DD': 4,'DD*':4,'F*':0, 'F': 0, 'I': 0}         
 # now making the overall sheet for all roll no
   with open('names-roll.csv', 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
          wb = Workbook()
          wb.create_sheet(index=0,title="Overall")
          sheet=wb["Overall"]
          sheet.append(["Roll No.",row['Roll']])
          sheet.append(["Name of Student ",row['Name']])
          a=row['Roll']
          sheet.append(["Discipline",a[4:6]])
          
          n=[]
          spi=[]
          credit_sem=[]
          total_credit=[]
          cpi=[]
          x=0
          w=0
          for j in range(1,11):
               v=(a,str(j))
               if v in data:
                     count=0
                     credit_obtain=0
                     for i in data[v]:
                       count+=int(i[1])
                       credit_obtain+=grades[i[3]]*(int(i[1]))
               if v in data:
                  n.append(j)   
                  credit_sem.append(count)     
                  spi.append(credit_obtain/count)
                  x+=count
                  total_credit.append(x)
                  w+=credit_obtain
                  cpi.append(w/x)
          sheet.append(["Semester No."]+[  x for x in n])
          sheet.append(["Semester wise Credit Taken"]+[x for x in credit_sem])
          sheet.append(["SPI"]+[round(x,2) for x in spi])
          sheet.append(["Total Credits Taken"]+[x for x in total_credit])
          sheet.append(["CPI"]+[round(x,2) for x in cpi])
          wb.remove(wb["Sheet"])
          wb.save(".\output\{}.xlsx".format(a))  
  # load the workbook create extra sheet and add the semester detail         
          wb=load_workbook(".\output\{}.xlsx".format(a))
          p = 0
          for j in range(1,11):
               v=(a,str(j))
               if v in data:
                    p=p+1
                    wb.create_sheet(index=p,title="Sem{}".format(str(j)))
                    wb["Sem{}".format(str(j))].append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                    sheet=wb["Sem{}".format(str(j))]
                    for i in data[v]:
                         x=subject_data[i[0]]
                         sheet.append([ sheet.max_row,i[0],x[0],x[1],x[2],i[2],i[3] ])
          wb.save(".\output\{}.xlsx".format(a))

   return  
          
generate_marksheet()
     







