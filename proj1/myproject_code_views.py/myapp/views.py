from django.shortcuts import render
from django.http import FileResponse
from django.core.files.storage import FileSystemStorage
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
from django.core.mail import message, send_mail,EmailMessage
from openpyxl.styles import Alignment
import shutil

# Create your views here.
response1 = {}
#default value of posititbe and negative is 5 and -1
def main(request):
    if request.method == 'POST':
        DIR = "sample_output/marksheet"
        if not os.path.exists(DIR): 
            os.makedirs(DIR)
        
        output_DIR10=r"./media"
        if os.path.exists(output_DIR10):
            shutil.rmtree(output_DIR10)

        p_marks=5
        n_marks=-1
        p_marks = float(request.POST['positive'])
        n_marks = float(request.POST['negative'])
        file1 = request.FILES['file1']
        file_save1 = FileSystemStorage()
        file_save1.save(file1.name,file1)
        file2 = request.FILES['file2']
        file_save2 = FileSystemStorage()
        file_save2.save(file2.name,file2)
#name_roll dictionary store name corresponding to roll 
        name_roll={}
        with open('media/'+file1.name, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name_roll[row['roll']] = row['name']

        global response1
        response2={} 
#response1 store the data in dictionary form with key as roll and value as list similarly rwsponse 2
        with open('media/'+file2.name, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                response1[row['Roll Number']] = [row['Email address'],row['Score'],row['Name'],row['IITP webmail'],row['Phone (10 digit only)']]
                response2[row['Roll Number']] = [row['Timestamp'],row['Email address'],row['Score'],row['Name'],row['IITP webmail'],row['Phone (10 digit only)'],row['Roll Number']]
# answer_key store the answer_key key and roll_no_response store response of individual roll
        answer_key=[]
        roll_no_response={}
        with open('media/'+file2.name, 'r') as file:
             reader = csv.reader(file)
             for row in reader:
                roll_no_response[row[6]] = row[7:]
                if(row[6]=='ANSWER'):
                    answer_key = row[7:]

        roll_no_summation={}
        for key in response2:
            if key not in 'Roll Number':
                wb=Workbook()
                sheet=wb.active
                sheet.title = "quiz"
                wb.save(f'sample_output\\marksheet\\{key}.xlsx')
        ques_count = len(answer_key)
        for key in roll_no_response:
            Right = 0
            wrong = 0
            unattempted = 0
            for x in range(ques_count):
                if len(roll_no_response[key][x])==0:
                    unattempted+=1
                elif roll_no_response[key][x]!=answer_key[x]:
                    wrong+=1
                else:
                    Right+=1
            roll_no_summation[key] = [Right,wrong,unattempted]
#list_of_absent store the roll no which present in master csv but not in resposes csv
        list_of_absent=[]
        for key in name_roll:
            if key not in response1.keys():
                list_of_absent.append(key)
        
        for key in response2:
            wb=load_workbook(r'sample_output\\marksheet\\{}.xlsx'.format(key))
            ws = wb.worksheets[0]
            ws.merge_cells('A5:G5')
            ws['A5']="MARKSHEET:"
            ws["A5"].alignment = Alignment(horizontal="center")
            ws['A5'].font = Font(bold=True)
            ws['A6'] = "Name:"

            if key in name_roll.keys():
                ws['B6'] = name_roll[key]
                ws['B6'].font = Font(bold=True)
                ws['D6'] = "Exam:"
                ws['E6'] = "quiz"
                ws['E6'].font = Font(bold=True)
                ws['A7'] = "Roll Number:"
                ws['B7'] = key
                ws['B7'].font = Font(bold=True)
                heading = ['Right','Wrong','Not Attempt','Max']
                green=Font(color='008000')
                red=Font(color='f44336')
# setup the heading of marksheet
                for i,value in enumerate(heading):
                    ws.cell(column=i+2,row=9,value=value)
                    temp=get_column_letter(i+2)+'9'
                    ws[temp].font = Font(bold=True)
                number = ['No.',roll_no_summation[key][0],roll_no_summation[key][1],roll_no_summation[key][2],ques_count]
               
# overall count i.e right,wrong ,not attempt,max
                for i,value in enumerate(number):
                    ws.cell(column=i+1,row=10,value=value)
                    if i==1:
                        ws.cell(column=i+1,row=10).font=green
                    elif i==2:
                        ws.cell(column=i+1,row=10).font=red

                marking = ['Marking',p_marks,n_marks,0]
                for i,value in enumerate(marking):
                    ws.cell(column=i+1,row=11,value=value)
                    if i==1:
                        ws.cell(column=i+1,row=11).font=green
                    elif i==2:
                        ws.cell(column=i+1,row=11).font=red
#calculating and storing in a list finall marks after including negative marking
                total_mark=str(p_marks*roll_no_summation[key][0]+n_marks*roll_no_summation[key][1])+'/'+str(p_marks*ques_count)
                total = ['Total',p_marks*roll_no_summation[key][0],n_marks*roll_no_summation[key][1],0,total_mark]
                for i,value in enumerate(total):
                    ws.cell(column=i+1,row=12,value=value)
                    if i==1:
                        ws.cell(column=i+1,row=12).font=green
                    elif i==2:
                        ws.cell(column=i+1,row=12).font=red
                ws['A15'] = 'Student Ans'
                ws['A15'].font = Font(bold=True)
                ws['B15'] = 'Right Ans'
                ws['B15'].font = Font(bold=True)
                ws['E12'].font = Font(color=colors.BLUE)
                img = openpyxl.drawing.image.Image('iitp logo.png')
                img.anchor = 'A1'
                img.height = 75
                img.width = 700
                ws.add_image(img)
                for i,value in enumerate(roll_no_response[key]):
                    ws.cell(column=1,row=16+i,value=value)
                    cell = 'A'+str(i+16)
                    if roll_no_response[key][i] == answer_key[i]:
                        ws[cell].font = green
                    else:
                        ws[cell].font = red
                for i,value in enumerate(answer_key):
                    ws.cell(column=2,row=16+i,value=value)
                    cell = 'B'+str(16+i)
                    ws[cell].font = Font(color=colors.BLUE)
                wb.save(f'sample_output\\marksheet\\{key}.xlsx')
#creting the concise_sheet
        wb=Workbook()
        sheet=wb.active
        sheet.title = "concise_sheet"
        heading = ['Timestamp','Email address','Goggle_Score','Name','IITP webmail','Phone (10 digit only)','Score_Ater_negative','Roll Number']
#setting the heading of concise sheet
        for i in range(ques_count):
                heading.append('Unnamed: '+str(9+i))
        heading.append('statusAns')
        sheet.append(heading)
#now creating the overall concise response of each roll no after negative and save it to concise sheet
        for key in roll_no_response:
                if key in name_roll.keys():
                    total_mark=str(p_marks*roll_no_summation[key][0]+n_marks*roll_no_summation[key][1])+'/'+str(p_marks*ques_count)
                    list1 = [response2[key][0],response2[key][1],response2[key][2],response2[key][3],response2[key][4],response2[key][5],
                         total_mark,response2[key][6]]
                    for x in roll_no_response[key]:
                        list1.append(x)
                    temp = '['+str(roll_no_summation[key][0])+','+str(roll_no_summation[key][1])+','+str(roll_no_summation[key][2])+']'
                    list1.append(temp)
                    sheet.append(list1)
        wb.save(f'sample_output\\marksheet\\concise_sheet.xlsx')

        wb=load_workbook(r'sample_output\\marksheet\\concise_sheet.xlsx')
        sheet = wb.active
        for x in list_of_absent:
                sheet.append(['ABSENT','ABSENT','ABSENT',name_roll[x],'ABSENT','ABSENT','ABSENT',x])
        wb.save(f'sample_output\\marksheet\\concise_sheet.xlsx')
                
                
        return render(request,'myapp/index.html',{'message1':'Excel files has created'})
        
            

    return render(request,'myapp/index.html')

def send_emails(request):
    for key in response1:
        mail = EmailMessage("quiz marks", "PFA", "gkpamitmax@gmail.com", [response1[key][0],response1[key][3]])
        mail.attach_file(f'sample_output\\marksheet\\{key}.xlsx')
        mail.send(fail_silently=True)
    return render(request,'myapp/index.html',{'message2':'Emails are being sent'})

def concise_sheet(request):
    
    files = open('sample_output\\marksheet\\concise_sheet.xlsx','rb')
    response = FileResponse(files)
    return response
    return render(request,'myapp/index.html',{'message3':'Concise sheet is generated'})