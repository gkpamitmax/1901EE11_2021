# Importing Required Libraries
import csv
import os
import shutil
from typing import ValuesView
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from openpyxl.reader.excel import load_workbook
import re
os.system('cls')

feedback_data={}
course_data={}
student_data={}
course_register={}

def check_feedback():
     wb=openpyxl.load_workbook('./course_feedback_remaining.xlsx')
     sheet=wb['Sheet1']
     sheet.delete_rows(2, sheet.max_row+1)
     with open('course_registered_by_all_students.csv','r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            if row['subno'] in course_data:
              for w in course_data[row['subno']]:
                q=(row['rollno'],row['subno'],w)
                if q in feedback_data:
                  pass
                else:
                   if row['rollno'] in student_data:
                     o=[row['rollno'],int(row['register_sem']),int(row['schedule_sem']),row['subno']]+student_data[row['rollno'] ]
                     sheet.append(o)
                   else:
                      o=[row['rollno'],int(row['register_sem']),int(row['schedule_sem']),row['subno'],'N/A','N/A','N/A','N/A']
                      sheet.append(o)
     wb.save('./course_feedback_remaining.xlsx')
     file.close()



def feedback_not_submitted():
  

   with open('course_feedback_submitted_by_students.csv','r') as file:
     reader = csv.DictReader(file)
     for row in reader:
       	x=(row['stud_roll'],row['course_code'],row['feedback_type'])
        feedback_data[x]=1
   file.close()

   regex = r"([0-9.]+)-([0-9.]+)-([0-9.]+)"

   with open('course_master_dont_open_in_excel.csv','r') as file:
     reader = csv.DictReader(file)
     for row in reader:
        match=re.search(regex,row['ltp'])
        w=[]
        if(match[1]!='0'):
          w=[1]
        if(match[2]!='0'):
           w.append(2)
        if(match[3]!='0'):
           w.append(3)
        if(w):
          course_data[row['subno']]=w

   file.close()

   # for x in course_data:
	#    print(course_data[x])

   with open('studentinfo.csv', 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
         student_data[row['Roll No']]=[row['Name'],row['email'],row['aemail'],int(row['contact'])]
   file.close()

   with open('course_registered_by_all_students.csv','r') as file:
     reader = csv.DictReader(file)
     for row in reader:
       	x=(row['rollno'],row['subno'])
        course_register[x]=[row['register_sem'],row['schedule_sem']]
   file.close()

   check_feedback()
	




	# ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
	# output_file_name = "course_feedback_remaining.xlsx" 




feedback_not_submitted()
