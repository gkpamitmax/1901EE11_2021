import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

def write_on_new_folder(val,data,ex_l,header):
    if val not in ex_l:
        ex_l.append(data[2])
        wb=Workbook()
        sheet=wb.active
        sheet.append(header)
        sheet.append(data)
        wb.save(f'output_by_subject\\{data[2]}.xlsx')
    else:
        wb=load_workbook(r'output_by_subject\\{}.xlsx'.format(data[2]))
        sheet=wb.active
        sheet.append(data)
        wb.save(f'output_by_subject\\{data[2]}.xlsx')

 
def output_by_subject():
    new_folder = "output_by_subject"
    header = ["rollno","register_sem","subno","sub_type"]  
    sub_list = []

    if not os.path.exists(new_folder): 
        os.makedirs(new_folder)
    with open('regtable_old.csv','r') as file:
        stud_list=csv.reader(file)
        for data1 in stud_list:
            data= []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[2] =="subno"):continue
            write_on_new_folder(data[2],data,sub_list,header)
             
    return
    
def write_on_new_folder_roll(val,data,exl,header):
    if val not in exl:
        exl.append(data[0])
        wb=Workbook()
        sheet=wb.active
        sheet.append(header)
        sheet.append(data)
        wb.save(f'output_individual_roll\\{data[0]}.xlsx')
    else:
        wb=load_workbook(r'output_individual_roll\\{}.xlsx'.format(data[0]))
        sheet=wb.active
        sheet.append(data)
        wb.save(f'output_individual_roll\\{data[0]}.xlsx')

 
def output_individual_roll():
    new_folder = "output_individual_roll"
    header = ["rollno","register_sem","subno","sub_type"]
    roll_list = []
    
    if not os.path.exists(new_folder): 
        os.makedirs(new_folder)
    with open('regtable_old.csv','r') as file:
        stud_list=csv.reader(file)
        for data1 in stud_list:
            data= []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[0] =="rollno"):continue
            write_on_new_folder_roll(data[0],data,roll_list,header)
    return
 
output_by_subject()
output_individual_roll()